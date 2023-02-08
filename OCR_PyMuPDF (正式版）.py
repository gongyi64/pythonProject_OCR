#PyMuPDF1.19.0で。最新１．１９．３？だと動かない

from turtle import clear
import fitz
import openpyxl as px
from openpyxl.styles import Alignment
import pandas as pd

import numpy as np

import re

item_list =[]

filename = 'IMG_8822変更.pdf'

#filename = '委託業務実施管理チェック表.pdf'

# filename = r'C:\Users\406239\Desktop\python\200101_給与.pdf'

doc =fitz.open(filename)

# プログラム4｜PDFを1ページずつテキストを取得
for page in range(len(doc)):
    textblocks = doc[page].get_text('blocks')
    for textblock in textblocks:
        if textblock[4].isspace() == False:
            item_list.append([page,textblock[4]])

print('読み込んだRAWデータ')
print(item_list)


df_item= pd.DataFrame(item_list)
print(df_item)



df_dict = {}
for name, group in df_item.groupby(0):
    df_dict[name] = group


# print('分割したDF?')
# print(df_dict)


for i in range(10):

  df_dict[i]=df_dict[i].reset_index(drop=True)
  print(i,'番目のサブDF')
  print(df_dict[i])



# print('全体のDF')
# print(df_kinmu)


for i in range(10):
 #必要データ抽出　

  #氏名データ入りの行を抽出

 df_name= df_dict[i][4:5]

 # 勤務データ入りの行を抽出

 df_item = df_dict[i][9:40]



 print('必要データ抽出')

 print('氏名')

#  print(df_name)

 df_name[2]= df_name[1].str.split(pat='\n',expand=True)[0]

#  print(df_name)

 df_name=df_name.drop(columns=1)

#  print(df_name)

 #担当者のみ抽出（実施担当者：を区切り文字で分割）

 df_name['実施']= df_name[2].str.split('：',expand=True)[0]
 df_name['氏名']= df_name[2].str.split('：',expand=True)[1]

 df_name=df_name.drop(columns=2,)
 df_name=df_name.drop(columns='実施')

 #df_nameを担当者氏名のみに

 print(df_name)


#  print('勤務全体')

#  print(df_item)

 #各リストの要素を改行で分割して、別カラムのDFへ

 #df_item['担当']= df_name[0].str.split(pat='\n',expand=True)[0]

 df_item['日']= df_item[1].str.split(pat='\n',expand=True)[0]
 df_item['曜日']= df_item[1].str.split(pat='\n',expand=True)[1]
 df_item['業務内容']= df_item[1].str.split(pat='\n',expand=True)[2]



#  print(df_item)


 df_item=df_item.drop(columns=1)  #取り込み元データ列削除



 df_item['要員数']=1 #カウント用の列追加　全部１にしている。

 df_item=df_item.reset_index(drop=True)

 df_kinmu = {}
 df_kinmu[i]= df_item

 df=[]
 df=pd.DataFrame(df_kinmu[i])


 print('df')
 print(df)

 added_row =df_name.iloc[0]

 nissuu=len(df_item)-1

 print(nissuu)

 for j in range(nissuu):
  df_name = df_name.append(added_row)

  df_name=df_name.reset_index(drop=True)


 print(df_name)
 if i == 0:
  df0=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df0)

 if i == 1:
  df1=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df1)


 if i == 2:
  df2=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df2)

 if i == 3:
  df3=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df3)

 if i == 4:
  df4=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df4)

 if i == 5:
  df5=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df5)

 if i == 6:
  df6=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df6)
 if i == 7:
  df7=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df7)

 if i == 8:
  df8=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df8)

 if i == 9:
  df9=pd.concat([df_item,df_name],axis=1)
  print('DF',i)
  print(df9)

   # df1[i] = df1

df_concat= pd.concat([df0,df1,df2,df3,df4,df5,df6,df7,df8,df9])

df_concat=df_concat.reset_index(drop=True)

print(df_concat)

#日を数値型に変換（object→int64)

df_concat['日']=df_concat['日'].astype('uint')


print('日付でソート')

df_sort=df_concat.sort_values(by='日',ascending=True)

# df_sort=df_sort.reset_index(drop=True)

print(df_sort)

df_sort.to_excel('NT全体報告.xlsx')

df_sort_cut=df_sort.drop(columns ='氏名')

df_sort_cut=df_sort_cut.drop(columns = 0)

df_sort_cut=df_sort_cut.reset_index(drop=True)

print(df_sort_cut)

df_houkoku = df_sort_cut.pivot_table(index = ['業務内容'],columns = ['日','曜日'],values=['要員数'],aggfunc='sum')

print('列入れ替え')

df_houkoku.reset_index(drop=True)

print(df_houkoku)


#sort_order = ['運行業務（早出Ｓ）平日','運行業務（早出Ｓ）土・日・祝','運行業務（日勤ＴＤ）','運行業務（ニュース送出NT）','運行業務（遅出ＮＬ）','運行業務（宿泊）平日','運行業務（宿泊）土、日、祝','運行業務（宿明）','運行業務（保守・整備ほか）','緊急スタンバイ・ヘリ','制作技術業務','報道対応業務','休日','ＮＴ社内業務（研修）','ＮＴ社内業務（管理業務）']

sort_order = ['S勤務','S勤務（土日祝）','TD業務','NT業務','NL業務','宿泊業務','宿明業務','運行業務（保守・整備)','休日']

df_sortkey=pd.DataFrame({'業務内容':['NL業務','NT業務','S勤務','S勤務（土日祝）','TD業務 ','休日','宿明業務','宿泊業務','運行業務（保守・整備）'],
'sortkey':[5,4,1,2,3,9,7,6,8]})
print('ソートキーの列')
print(df_sortkey)

df_houkoku['sort_key']=[5,4,1,2,3,9,7,6,8]


print('sort前のDF')

print(df_houkoku)

df_houkoku = df_houkoku.sort_values('sort_key')

print('sortされたDF')
print(df_houkoku)



df_houkoku.drop('sort_key',axis=1,inplace=True)

print('最終報告データ')
print(df_houkoku)

df_houkoku.to_excel('NT報告.xlsx')

"""
#-------以下　EXCELの表の体裁調整−−−−−

import openpyxl
from openpyxl.styles.borders import Border , Side
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment

# book = openpyxl.load_workbook('202109_5512沖縄管内_沖縄事業所_委託入力チェック.xlsx')

filepath = '{0}_pandas_to_excel.xlsx'.format(num)

print(filepath)

book = openpyxl.load_workbook(filepath)

sheet = book['Sheet1']

# SheetRowNo = 2

# sheet.insert_rows(SheetRowNo)


#全行の高さを75にする。
# sheet.row_dimensions[1].height = 15

for i in range(1,3):
     sheet.row_dimensions[i].height = 37.6

for i in range(3,17):
     sheet.row_dimensions[i].height = 75.2

# sheet.merge_cells('B1:B2')
#sheet.merge_cells('A1:A2')



#XLwings  版　インストール未
# import xlwings as xw
# wb = xw.Book("202110_5512沖縄管内_沖縄事業所_委託入力チェック.xlsx")
# wb.sheets[0]['1,15'].api.RowHeight =15

#--------------------------------------
# for i in range(3,17):
#          sheet.column_dimensions[i].width = 30

sheet.column_dimensions['A'].width = 32
sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 10
sheet.column_dimensions['D'].width = 10
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 10
sheet.column_dimensions['G'].width = 10
sheet.column_dimensions['H'].width = 10
sheet.column_dimensions['I'].width = 10
sheet.column_dimensions['J'].width = 10
sheet.column_dimensions['K'].width = 10
sheet.column_dimensions['L'].width = 10
sheet.column_dimensions['M'].width = 10
sheet.column_dimensions['N'].width = 10
sheet.column_dimensions['O'].width = 10
sheet.column_dimensions['P'].width = 10
sheet.column_dimensions['Q'].width = 10
sheet.column_dimensions['R'].width = 10

#---------------------------------------------
#日付表示の変更


for i in range(2,18):

      sheet.cell(row=i,column=2).font =openpyxl.styles.fonts.Font(color = '2B2B2B',size=14)



for i in range(2,19):

      sheet.cell(row=1,column=i).font =openpyxl.styles.fonts.Font(color = '2B2B2B',size=12)


for i in range(3,18):
     for j in range(2,19):
      sheet.cell(row=i,column=j).font =openpyxl.styles.fonts.Font(color = 'FF0000',size=30)

     #  sheet[col].font =openpyxl.styles.fonts.Font(color = 'FF0000',size=40)



#セル幅の自動設定

# set column width
# for col in sheet.columns:
#     max_length = 0
#     column = col[0].column_letter

#     for cell in col:
#         if len(str(cell.value)) > max_length:
#             max_length = len(str(cell.value))

#     adjusted_width = (max_length + 2) * 1.7
#     sheet.column_dimensions[column].width = adjusted_width

#openpyxl 3以降でcolumn_dimensionsの添字が列番号の数値から列名の文字列に変更されたから
#--------------------------------------------------

#文字そろえ
for row in sheet:
     for cell in row:


      if cell.row ==1:
           cell.alignment = Alignment(horizontal = 'center',vertical ='center')

for row in sheet["A1:A18"]:
     for cell in row:

           cell.alignment = Alignment(horizontal = 'center',vertical ='center')

for row in sheet["B2:S17"]:
     for cell in row:

           cell.alignment = Alignment(horizontal = 'center',vertical ='center')



side1= Side(style = 'thick',color='000000')
border_aro = Border(top=side1,bottom=side1,left=side1,right=side1)

side2= Side(style = 'dotted',color ='000000')

# border_b = Border(bottom =side2)

# sheet['C1:R1'].border = border_b

for row_num in sheet['B1:R1']:

        for cell in row_num:
             cell.border = Border(bottom= side2)

# for row_num in sheet['C2:R2']:

#         for cell in row_num:
#              cell.border = Border(top= side2)


for row_num in sheet['A3:R16']:

        for cell in row_num:
             cell.border = Border(left = side1,right = side1,top = side1,bottom= side1)



for row_num in sheet['A1:A12']:

        for cell in row_num:
             cell.fill = PatternFill(patternType ='solid',start_color = 'D1FE7B',end_color = 'D1FE7B')

for row_num in sheet['A13:A14']:

        for cell in row_num:
             cell.fill = PatternFill(patternType ='solid',start_color = 'FF69B4',end_color = 'FF69B4')


for row_num in sheet['A15:A16']:

        for cell in row_num:
             cell.fill = PatternFill(patternType ='solid',start_color = '808080',end_color = '808080')



book.save('{0}_test_nt.xlsx'.format(num))

"""